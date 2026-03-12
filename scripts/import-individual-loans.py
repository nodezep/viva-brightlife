import os
import sys
import json
from datetime import datetime, date

import openpyxl
import requests


def load_env_from_file(path: str):
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


def to_iso_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value).date().isoformat()
        except ValueError:
            pass
    raise ValueError(f"Unsupported date value: {value!r}")


def normalize_rate(raw):
    if raw is None or raw == "":
        return 0.0, 0.0
    rate = float(raw)
    # If input is 0.2, treat as 20%. If input is 20, treat as 20%.
    if rate <= 1:
        return rate * 100, rate
    return rate, rate / 100


def main():
    excel_path = (
        sys.argv[1]
        if len(sys.argv) > 1
        else r"C:\Users\micha\Desktop\LOANSMANUAL.xlsx"
    )
    dry_run = "--dry-run" in sys.argv

    load_env_from_file(".env.local")
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_ANON_KEY")

    if not url or not key:
        print("Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env.local", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if row and any(cell == "S/NO" for cell in row):
            header_row = row
            break

    if not header_row:
        print("Could not find header row with 'S/NO'.", file=sys.stderr)
        sys.exit(1)

    def normalize_header(value: str) -> str:
        return " ".join(value.replace("\n", " ").split()).upper()

    headers = {
        normalize_header(str(value)): idx
        for idx, value in enumerate(header_row)
        if value
    }
    required = [
        "S/NO",
        "JINA",
        "KIASI CHA MKOPO",
        "TAREHE YA MKOPAJI",
        "IDADI YA SIKU ZA MALIMBIKIZO",
        "ASILIMIA YA RIBA",
        "RIBA",
        "MUDA WA MKOPO(MWEZI)",
        "MALIPO YA MKOPO",
        "MKOPO+REJESHO",
        "DENI",
        "NAMBA YA SIMU"
    ]
    required_keys = [normalize_header(name) for name in required]
    missing = [name for name, key in zip(required, required_keys) if key not in headers]
    if missing:
        print(f"Missing expected columns: {missing}", file=sys.stderr)
        sys.exit(1)

    session = requests.Session()
    session.headers.update(
        {
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation,resolution=merge-duplicates",
        }
    )

    inserted = 0
    skipped = 0
    errors = 0

    start_row = ws.min_row + header_row.index("S/NO")  # header row position
    # Find actual header row index
    header_index = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        if row and any(cell == "S/NO" for cell in row):
            header_index = idx
            break

    for row in ws.iter_rows(min_row=(header_index or 1) + 1, values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        try:
            serial = str(row[headers[normalize_header("S/NO")]]).strip()
            name = str(row[headers[normalize_header("JINA")]]).strip()
            principal = float(row[headers[normalize_header("KIASI CHA MKOPO")]] or 0)
            disbursement_date = to_iso_date(row[headers[normalize_header("TAREHE YA MKOPAJI")]])
            months = int(row[headers[normalize_header("MUDA WA MKOPO(MWEZI)")]] or 1)
            phone = row[headers[normalize_header("NAMBA YA SIMU")]]
            phone = str(phone).strip() if phone is not None else None

            days_overdue = row[headers[normalize_header("IDADI YA SIKU ZA MALIMBIKIZO")]]
            days_overdue = int(days_overdue) if days_overdue is not None else 0

            interest_raw = row[headers[normalize_header("ASILIMIA YA RIBA")]]
            interest_rate_percent, interest_rate_decimal = normalize_rate(interest_raw)

            riba = row[headers[normalize_header("RIBA")]]
            riba = float(riba) if riba is not None else principal * interest_rate_decimal

            total = row[headers[normalize_header("MKOPO+REJESHO")]]
            total = float(total) if total is not None else principal + riba

            amount_paid = row[headers[normalize_header("MALIPO YA MKOPO")]]
            amount_paid = float(amount_paid) if amount_paid is not None else 0.0

            balance = row[headers[normalize_header("DENI")]]
            balance = float(balance) if balance is not None else total - amount_paid

            member_number = serial
            loan_number = f"BIN-{serial}"

            if not name or not disbursement_date:
                skipped += 1
                continue

            if dry_run:
                inserted += 1
                continue

            member_payload = {
                "member_number": member_number,
                "full_name": name,
                "phone": phone or None
            }
            member_resp = session.post(
                f"{url}/rest/v1/members?on_conflict=member_number",
                data=json.dumps(member_payload),
            )
            member_resp.raise_for_status()
            member_data = member_resp.json()
            if not member_data:
                raise RuntimeError("Member insert/upsert returned empty response.")
            member_id = member_data[0]["id"]

            loan_payload = {
                "loan_number": loan_number,
                "member_id": member_id,
                "loan_type": "binafsi",
                "principal_amount": principal,
                "disbursement_date": disbursement_date,
                "security_amount": riba,
                "cycle_count": months,
                "weekly_installment": total,
                "monthly_installment": 0,
                "amount_withdrawn": amount_paid,
                "outstanding_balance": balance,
                "overdue_amount": 0,
                "interest_rate": interest_rate_percent,
                "duration_months": months,
                "days_overdue": days_overdue,
                "status": "active"
            }

            loan_resp = session.post(
                f"{url}/rest/v1/loans?on_conflict=loan_number",
                data=json.dumps(loan_payload),
            )
            loan_resp.raise_for_status()
            inserted += 1

        except Exception as exc:
            errors += 1
            print(f"Row error: {exc}", file=sys.stderr)

    print(f"Inserted: {inserted}, Skipped: {skipped}, Errors: {errors}")


if __name__ == "__main__":
    main()
